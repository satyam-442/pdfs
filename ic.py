import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# Step 1: Read Excel file
file_path = "itsm_raw.xlsx"
output_path = "itsm-final.xlsx"

df = pd.read_excel(file_path)
print("Column Headers:")
print(df.columns.tolist())

wb = load_workbook(file_path)
ws = wb.active

# -------------------------
#  FIND IMPORTANT COLUMNS
# -------------------------
for col in range(1, ws.max_column + 1):
    header = ws.cell(row=1, column=col).value

    if header == "Created":
        created_col = col     # Example: E column

    if header == "Number":    # Incident column (Column B)
        incident_col = col

# -------------------------
#  ADD AGE COLUMNS
# -------------------------
age_col = created_col + 1         # Age ("__ Days")
age_num_col = created_col + 2     # Age Number (just numeric days)

ws.insert_cols(age_col)
ws.insert_cols(age_num_col)

ws.cell(row=1, column=age_col).value = "Age"
ws.cell(row=1, column=age_num_col).value = "Age Number"

# Insert formulas
for row in range(2, ws.max_row + 1):
    created_cell = ws.cell(row=row, column=created_col).coordinate

    ws.cell(row=row, column=age_col).value = (
        f'=CONCATENATE(CEILING(TODAY() - {created_cell}, 1), " Days")'
    )

    ws.cell(row=row, column=age_num_col).value = (
        f'=CEILING(TODAY() - {created_cell}, 1)'
    )

# -------------------------
#  CONDITIONAL FORMATTING 
#  Using your formula:
#   =DAYS(TODAY(), $E2) > 14
# -------------------------

red_font = Font(color="FF0000")
formula = "=DAYS(TODAY(),$E2) > 14"
rule = FormulaRule(formula=[formula], stopIfTrue=True, font=red_font)

last_row = ws.max_row

# Convert column numbers → Excel letters
incident_letter = get_column_letter(incident_col)
age_letter = get_column_letter(age_col)
age_num_letter = get_column_letter(age_num_col)

# Build valid range strings like "B2:B200"
incident_range = f"{incident_letter}2:{incident_letter}{last_row}"
age_range = f"{age_letter}2:{age_letter}{last_row}"
age_num_range = f"{age_num_letter}2:{age_num_letter}{last_row}"

# Apply conditional formatting
ws.conditional_formatting.add(incident_range, rule)
ws.conditional_formatting.add(age_range, rule)
ws.conditional_formatting.add(age_num_range, rule)

# Save final Excel file
wb.save(output_path)

print("DONE! Age columns added + conditional formatting applied.")
